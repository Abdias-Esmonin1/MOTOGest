from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from django.http import HttpResponse
from apps.paiements.models import Paiement
from apps.pannes.models import Panne
from apps.finances.models import Depense

def export_paiements_excel(date_debut=None, date_fin=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Paiements"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1a472a")

    headers = ["Date", "Moto", "Montant attendu (FCFA)", "Montant versé (FCFA)", "Statut", "Observations"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    qs = Paiement.objects.select_related('moto').order_by('-date')
    if date_debut:
        qs = qs.filter(date__gte=date_debut)
    if date_fin:
        qs = qs.filter(date__lte=date_fin)

    for row, p in enumerate(qs, 2):
        ws.cell(row=row, column=1, value=str(p.date))
        ws.cell(row=row, column=2, value=str(p.moto))
        ws.cell(row=row, column=3, value=int(p.montant_attendu))
        ws.cell(row=row, column=4, value=int(p.montant_verse))
        ws.cell(row=row, column=5, value=p.get_statut_display())
        ws.cell(row=row, column=6, value=p.observations)

    for col in ws.columns:
        max_len = max(len(str(c.value or '')) for c in col) + 4
        ws.column_dimensions[col[0].column_letter].width = min(max_len, 40)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="paiements.xlsx"'
    wb.save(response)
    return response
